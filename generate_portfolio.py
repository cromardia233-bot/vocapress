#!/usr/bin/env python3
"""
포트폴리오 대시보드 HTML 생성 스크립트
엑셀 데이터를 읽어 모바일 친화적인 단일 HTML 파일로 변환
"""

import openpyxl
import json
import datetime
import os

EXCEL_PATH = '/Users/suhunchoi/Downloads/00. Portfolio.xlsx'
OUTPUT_PATH = '/Users/suhunchoi/Desktop/Claude code/portfolio.html'


def safe_val(v):
    """값을 JSON 직렬화 가능한 형태로 변환"""
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (v != v):  # NaN 체크
            return None
        return v
    return str(v)


def read_holdings(wb):
    """보유종목 데이터 읽기"""
    ws = wb['Holdings']
    holdings = []
    summary = {}

    # 총자산
    total_assets = ws.cell(row=1, column=11).value

    for row in ws.iter_rows(min_row=4, max_row=100, values_only=True):
        vals = list(row[:22])
        top_pick = vals[0]
        market = vals[1]
        name = vals[2]
        ticker = vals[3]
        tag = vals[4]
        industry = vals[5]
        weight = vals[6]
        buy_price = vals[7]
        current_price = vals[8]
        shares = vals[9]
        current_value = vals[10]
        pnl = vals[11]
        return_pct = vals[12]
        daily_change = vals[13]
        daily_pnl = vals[14]

        if name is None:
            continue

        # 예수금 행
        if name == '예수금':
            summary['cash'] = safe_val(current_value)
            continue

        # 요약 행 (미국주식, 호주주식, 한국주식)
        if market and name == market:
            summary[market] = {
                'weight': safe_val(weight),
                'value': safe_val(current_value),
                'pnl': safe_val(pnl),
                'return': safe_val(return_pct)
            }
            continue

        # 환율 행
        if market and 'KRW' in str(market):
            continue

        holdings.append({
            'rank': safe_val(top_pick),
            'market': safe_val(market),
            'name': safe_val(name),
            'ticker': safe_val(ticker),
            'tag': safe_val(tag),
            'industry': safe_val(industry),
            'weight': safe_val(weight),
            'buyPrice': safe_val(buy_price),
            'currentPrice': safe_val(current_price),
            'shares': safe_val(shares),
            'currentValue': safe_val(current_value),
            'pnl': safe_val(pnl),
            'returnPct': safe_val(return_pct),
            'dailyChange': safe_val(daily_change),
            'dailyPnl': safe_val(daily_pnl),
        })

    return {
        'totalAssets': safe_val(total_assets),
        'holdings': holdings,
        'summary': summary
    }


def read_portfolio_value(wb):
    """월별 포트폴리오 가치 데이터 읽기"""
    ws = wb['Portfolio Value']
    months = []

    for row in ws.iter_rows(min_row=5, max_row=20, values_only=True):
        vals = list(row[:17])
        month_num = vals[1]
        if month_num is None or not isinstance(month_num, (int, float)):
            continue

        end_date = vals[3]
        ending_value = vals[9]
        portfolio_return = vals[11]
        portfolio_indexed = vals[12]
        sp500_return = vals[13]
        sp500_indexed = vals[14]
        kospi_return = vals[15]
        kospi_indexed = vals[16]

        # 유효한 데이터만 (ending_value가 있고 portfolio_return이 숫자인 경우)
        if ending_value and isinstance(portfolio_return, (int, float)):
            if isinstance(portfolio_return, str):
                continue
            months.append({
                'month': safe_val(month_num),
                'date': safe_val(end_date),
                'endValue': safe_val(ending_value),
                'portfolioReturn': safe_val(portfolio_return),
                'portfolioIndexed': safe_val(portfolio_indexed),
                'sp500Return': safe_val(sp500_return),
                'sp500Indexed': safe_val(sp500_indexed),
                'kospiReturn': safe_val(kospi_return),
                'kospiIndexed': safe_val(kospi_indexed),
            })

    # 시작 가치 (첫달 starting value)
    starting_value = ws.cell(row=5, column=6).value

    return {
        'startingValue': safe_val(starting_value),
        'months': months
    }


def read_investment_judgement(wb):
    """투자전략 및 종목별 판단 기록 읽기"""
    ws = wb['Investment Judgement']

    # 자산배분 전략
    strategy_date = safe_val(ws.cell(row=4, column=1).value)
    asset_allocation = str(ws.cell(row=4, column=8).value or '')
    investment_strategy = str(ws.cell(row=4, column=9).value or '')

    # 섹터별 비중/목표수익률
    sectors = []
    sector_names = ['반도체', '휴머노이드', '원자재', '코어주식']
    sector_cols = [8, 9, 10, 11]
    for i, (name, col) in enumerate(zip(sector_names, sector_cols)):
        weight = ws.cell(row=6, column=col).value
        target_return = ws.cell(row=7, column=col).value
        sectors.append({
            'name': name,
            'weight': safe_val(weight),
            'targetReturn': safe_val(target_return)
        })

    # 투자전략 메모 (행 11)
    strategy_memo = safe_val(ws.cell(row=11, column=8).value)

    # 종목별 판단 기록 (행 19~)
    judgements = []
    for row in ws.iter_rows(min_row=19, max_row=60, values_only=True):
        vals = list(row[:15])
        date = vals[0]
        category = vals[1]
        name = vals[2]
        ticker = vals[3]
        tag = vals[4]
        industry = vals[5]
        decision = vals[6]
        thesis = vals[7]
        logic = vals[8]
        result = vals[12]
        result_return = vals[13]

        if name is None and date is None:
            continue
        if name is None:
            continue

        judgements.append({
            'date': safe_val(date),
            'category': safe_val(category),
            'name': safe_val(name),
            'ticker': safe_val(ticker),
            'tag': safe_val(tag),
            'industry': safe_val(industry),
            'decision': safe_val(decision),
            'thesis': safe_val(thesis),
            'logic': safe_val(logic),
            'result': safe_val(result),
            'resultReturn': safe_val(result_return),
        })

    return {
        'strategyDate': strategy_date,
        'assetAllocation': asset_allocation,
        'investmentStrategy': investment_strategy,
        'sectors': sectors,
        'strategyMemo': strategy_memo,
        'judgements': judgements
    }


def read_watchlist(wb):
    """워치리스트 데이터 읽기"""
    ws = wb['Watchlist']
    groups = {}
    current_group = '기타'

    for row in ws.iter_rows(min_row=4, max_row=100, values_only=True):
        vals = list(row[:12])
        source = vals[0]
        market = vals[1]
        name = vals[2]
        ticker = vals[3]
        tag = vals[4]
        industry = vals[5]
        daily_change = vals[6]
        current_price = vals[7]
        market_cap = vals[10]
        idea = vals[11]

        # 그룹 헤더 행 (경로만 있고 종목명 없음)
        if source and not name:
            current_group = source
            if current_group not in groups:
                groups[current_group] = []
            continue

        if name is None:
            continue

        group_key = source if source else current_group

        if group_key not in groups:
            groups[group_key] = []

        groups[group_key].append({
            'source': safe_val(source or group_key),
            'market': safe_val(market),
            'name': safe_val(name),
            'ticker': safe_val(ticker),
            'tag': safe_val(tag),
            'industry': safe_val(industry),
            'dailyChange': safe_val(daily_change),
            'currentPrice': safe_val(current_price),
            'marketCap': safe_val(market_cap),
            'idea': safe_val(idea),
        })

    return groups


def read_trades(wb):
    """거래내역 읽기"""
    ws = wb['Trades']
    trades = []

    for row in ws.iter_rows(min_row=4, max_row=20, values_only=True):
        vals = list(row[:15])
        date = vals[0]
        rationale = vals[1]
        action = vals[2]
        market = vals[3]
        name = vals[4]
        ticker = vals[5]
        tag = vals[6]
        industry = vals[7]
        weight = vals[8]
        buy_price = vals[9]
        sell_price = vals[10]
        shares = vals[11]
        current_value = vals[12]
        pnl = vals[13]
        return_pct = vals[14]

        if name is None:
            continue

        # action 정리
        if isinstance(action, (int, float)):
            action_str = f"{int(action)}차 매수"
        else:
            action_str = str(action) if action else ''

        trades.append({
            'date': safe_val(date),
            'rationale': safe_val(rationale),
            'action': action_str,
            'market': safe_val(market),
            'name': safe_val(name),
            'ticker': safe_val(ticker),
            'tag': safe_val(tag),
            'industry': safe_val(industry),
            'weight': safe_val(weight),
            'buyPrice': safe_val(buy_price),
            'sellPrice': safe_val(sell_price),
            'shares': safe_val(shares),
            'currentValue': safe_val(current_value),
            'pnl': safe_val(pnl),
            'returnPct': safe_val(return_pct),
        })

    return trades


def generate_html(data):
    """HTML 파일 생성"""
    data_json = json.dumps(data, ensure_ascii=False, indent=2)

    html = f'''<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
<title>Portfolio Dashboard</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
:root {{
  --bg: #0f1117;
  --card: #1a1d27;
  --card-hover: #22253a;
  --border: #2a2d3a;
  --text: #e4e6eb;
  --text-sub: #8b8fa3;
  --accent: #6c5ce7;
  --accent-light: #a29bfe;
  --green: #00b894;
  --red: #e17055;
  --blue: #74b9ff;
  --yellow: #fdcb6e;
  --orange: #e17055;
  --tab-bg: #1a1d27;
  --tab-active: #6c5ce7;
}}

body {{
  font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
  background: var(--bg);
  color: var(--text);
  min-height: 100vh;
  -webkit-font-smoothing: antialiased;
  padding-bottom: 70px;
}}

/* 헤더 */
.header {{
  background: linear-gradient(135deg, #1a1d27 0%, #2d1b69 100%);
  padding: 20px 16px 16px;
  position: sticky;
  top: 0;
  z-index: 100;
  border-bottom: 1px solid var(--border);
}}
.header h1 {{
  font-size: 18px;
  font-weight: 700;
  color: var(--text);
  display: flex;
  align-items: center;
  gap: 8px;
}}
.header .date {{
  font-size: 12px;
  color: var(--text-sub);
  margin-top: 4px;
}}

/* 탭 네비게이션 */
.tab-nav {{
  display: flex;
  background: var(--tab-bg);
  border-bottom: 1px solid var(--border);
  overflow-x: auto;
  -webkit-overflow-scrolling: touch;
  scrollbar-width: none;
  position: sticky;
  top: 68px;
  z-index: 99;
}}
.tab-nav::-webkit-scrollbar {{ display: none; }}
.tab-btn {{
  flex: 0 0 auto;
  padding: 12px 16px;
  font-size: 13px;
  font-weight: 600;
  color: var(--text-sub);
  background: none;
  border: none;
  border-bottom: 2px solid transparent;
  cursor: pointer;
  white-space: nowrap;
  transition: all 0.2s;
}}
.tab-btn.active {{
  color: var(--accent-light);
  border-bottom-color: var(--tab-active);
}}

/* 탭 컨텐츠 */
.tab-content {{ display: none; padding: 16px; }}
.tab-content.active {{ display: block; }}

/* 대시보드 요약 카드 */
.summary-grid {{
  display: grid;
  grid-template-columns: 1fr 1fr;
  gap: 12px;
  margin-bottom: 20px;
}}
.summary-card {{
  background: var(--card);
  border-radius: 12px;
  padding: 16px;
  border: 1px solid var(--border);
}}
.summary-card.full {{ grid-column: 1 / -1; }}
.summary-card .label {{
  font-size: 11px;
  color: var(--text-sub);
  text-transform: uppercase;
  letter-spacing: 0.5px;
  margin-bottom: 6px;
}}
.summary-card .value {{
  font-size: 22px;
  font-weight: 700;
}}
.summary-card .value.positive {{ color: var(--green); }}
.summary-card .value.negative {{ color: var(--red); }}
.summary-card .sub {{
  font-size: 12px;
  color: var(--text-sub);
  margin-top: 4px;
}}

/* 차트 컨테이너 */
.chart-container {{
  background: var(--card);
  border-radius: 12px;
  padding: 16px;
  border: 1px solid var(--border);
  margin-bottom: 16px;
}}
.chart-container h3 {{
  font-size: 14px;
  font-weight: 600;
  margin-bottom: 12px;
  color: var(--text);
}}
.chart-container canvas {{
  max-height: 250px;
}}

/* 보유종목 카드 */
.holding-card {{
  background: var(--card);
  border-radius: 12px;
  padding: 14px 16px;
  border: 1px solid var(--border);
  margin-bottom: 10px;
  transition: background 0.2s;
}}
.holding-card:active {{ background: var(--card-hover); }}
.holding-header {{
  display: flex;
  justify-content: space-between;
  align-items: flex-start;
  margin-bottom: 8px;
}}
.holding-name {{
  font-size: 14px;
  font-weight: 600;
  line-height: 1.3;
  max-width: 65%;
}}
.holding-return {{
  font-size: 16px;
  font-weight: 700;
  text-align: right;
}}
.holding-meta {{
  display: flex;
  gap: 8px;
  flex-wrap: wrap;
  margin-bottom: 6px;
}}
.tag {{
  font-size: 10px;
  padding: 2px 8px;
  border-radius: 4px;
  background: rgba(108, 92, 231, 0.15);
  color: var(--accent-light);
}}
.tag.market {{ background: rgba(116, 185, 255, 0.15); color: var(--blue); }}
.holding-details {{
  display: flex;
  justify-content: space-between;
  font-size: 12px;
  color: var(--text-sub);
}}
.holding-details span {{ display: flex; flex-direction: column; }}
.holding-details .detail-label {{ font-size: 10px; color: var(--text-sub); margin-bottom: 2px; }}
.holding-details .detail-value {{ font-size: 13px; color: var(--text); font-weight: 500; }}

/* 섹션 타이틀 */
.section-title {{
  font-size: 15px;
  font-weight: 700;
  margin: 20px 0 12px;
  padding-bottom: 8px;
  border-bottom: 1px solid var(--border);
  color: var(--text);
}}
.section-title:first-child {{ margin-top: 0; }}

/* 전략 카드 */
.strategy-card {{
  background: var(--card);
  border-radius: 12px;
  padding: 16px;
  border: 1px solid var(--border);
  margin-bottom: 12px;
}}
.strategy-card h4 {{
  font-size: 13px;
  font-weight: 600;
  color: var(--accent-light);
  margin-bottom: 8px;
}}
.strategy-card p, .strategy-card pre {{
  font-size: 13px;
  color: var(--text);
  line-height: 1.6;
  white-space: pre-wrap;
  word-break: break-word;
}}

/* 섹터 비중 바 */
.sector-bar {{
  display: flex;
  border-radius: 8px;
  overflow: hidden;
  height: 28px;
  margin: 12px 0;
}}
.sector-bar div {{
  display: flex;
  align-items: center;
  justify-content: center;
  font-size: 10px;
  font-weight: 600;
  color: #fff;
}}
.sector-legend {{
  display: flex;
  flex-wrap: wrap;
  gap: 12px;
  margin-bottom: 16px;
}}
.sector-legend-item {{
  display: flex;
  align-items: center;
  gap: 6px;
  font-size: 12px;
  color: var(--text-sub);
}}
.sector-legend-item .dot {{
  width: 10px;
  height: 10px;
  border-radius: 50%;
}}

/* 아코디언 */
.accordion {{ margin-bottom: 8px; }}
.accordion-header {{
  background: var(--card);
  border: 1px solid var(--border);
  border-radius: 10px;
  padding: 14px 16px;
  cursor: pointer;
  display: flex;
  justify-content: space-between;
  align-items: center;
  transition: background 0.2s;
}}
.accordion-header:active {{ background: var(--card-hover); }}
.accordion-header .acc-title {{
  font-size: 13px;
  font-weight: 600;
  flex: 1;
}}
.accordion-header .acc-date {{
  font-size: 11px;
  color: var(--text-sub);
  margin-right: 8px;
}}
.accordion-header .acc-decision {{
  font-size: 11px;
  padding: 2px 8px;
  border-radius: 4px;
  font-weight: 600;
  margin-right: 8px;
}}
.acc-decision.buy {{ background: rgba(0, 184, 148, 0.2); color: var(--green); }}
.acc-decision.sell {{ background: rgba(225, 112, 85, 0.2); color: var(--red); }}
.acc-decision.hold {{ background: rgba(253, 203, 110, 0.2); color: var(--yellow); }}
.acc-decision.notinvest {{ background: rgba(139, 143, 163, 0.2); color: var(--text-sub); }}
.accordion-header .arrow {{
  font-size: 12px;
  color: var(--text-sub);
  transition: transform 0.2s;
}}
.accordion.open .arrow {{ transform: rotate(180deg); }}
.accordion-body {{
  display: none;
  background: var(--card);
  border: 1px solid var(--border);
  border-top: none;
  border-radius: 0 0 10px 10px;
  padding: 14px 16px;
}}
.accordion.open .accordion-header {{ border-radius: 10px 10px 0 0; }}
.accordion.open .accordion-body {{ display: block; }}
.accordion-body .field {{
  margin-bottom: 10px;
}}
.accordion-body .field-label {{
  font-size: 11px;
  font-weight: 600;
  color: var(--accent-light);
  margin-bottom: 4px;
}}
.accordion-body .field-value {{
  font-size: 12px;
  color: var(--text);
  line-height: 1.6;
  white-space: pre-wrap;
  word-break: break-word;
}}

/* 워치리스트 그룹 */
.wl-group {{ margin-bottom: 20px; }}
.wl-group-title {{
  font-size: 14px;
  font-weight: 700;
  color: var(--accent-light);
  padding: 8px 0;
  border-bottom: 1px solid var(--border);
  margin-bottom: 10px;
}}
.wl-card {{
  background: var(--card);
  border-radius: 10px;
  padding: 12px 14px;
  border: 1px solid var(--border);
  margin-bottom: 8px;
}}
.wl-card-header {{
  display: flex;
  justify-content: space-between;
  align-items: center;
  margin-bottom: 6px;
}}
.wl-card-name {{ font-size: 13px; font-weight: 600; max-width: 70%; }}
.wl-card-change {{ font-size: 14px; font-weight: 700; }}
.wl-card-meta {{
  display: flex;
  gap: 6px;
  flex-wrap: wrap;
  font-size: 11px;
  color: var(--text-sub);
}}

/* 거래내역 */
.trade-card {{
  background: var(--card);
  border-radius: 10px;
  padding: 14px 16px;
  border: 1px solid var(--border);
  margin-bottom: 8px;
}}
.trade-header {{
  display: flex;
  justify-content: space-between;
  align-items: center;
  margin-bottom: 8px;
}}
.trade-name {{ font-size: 13px; font-weight: 600; }}
.trade-action {{
  font-size: 11px;
  padding: 3px 10px;
  border-radius: 4px;
  font-weight: 600;
}}
.trade-action.buy {{ background: rgba(0, 184, 148, 0.2); color: var(--green); }}
.trade-action.sell {{ background: rgba(225, 112, 85, 0.2); color: var(--red); }}
.trade-details {{
  display: grid;
  grid-template-columns: repeat(3, 1fr);
  gap: 8px;
  font-size: 12px;
}}
.trade-details .td-label {{ color: var(--text-sub); font-size: 10px; }}
.trade-details .td-value {{ color: var(--text); font-weight: 500; margin-top: 2px; }}
.trade-rationale {{
  margin-top: 8px;
  font-size: 12px;
  color: var(--text-sub);
  line-height: 1.5;
  border-top: 1px solid var(--border);
  padding-top: 8px;
}}

/* 자산배분 파이 레전드 */
.alloc-grid {{
  display: grid;
  grid-template-columns: 1fr 1fr;
  gap: 10px;
  margin-top: 12px;
}}
.alloc-item {{
  display: flex;
  align-items: center;
  gap: 8px;
  font-size: 13px;
}}
.alloc-dot {{
  width: 12px;
  height: 12px;
  border-radius: 50%;
  flex-shrink: 0;
}}
.alloc-info .alloc-label {{ color: var(--text-sub); font-size: 11px; }}
.alloc-info .alloc-value {{ font-weight: 600; }}

/* 시장별 그룹 헤더 */
.market-group-header {{
  font-size: 13px;
  font-weight: 700;
  color: var(--blue);
  padding: 10px 0 6px;
  margin-top: 8px;
  display: flex;
  justify-content: space-between;
  align-items: center;
  border-bottom: 1px solid var(--border);
  margin-bottom: 8px;
}}
.market-group-header .market-return {{
  font-size: 12px;
  font-weight: 600;
}}

/* 가격 새로고침 */
#refreshBtn.loading {{
  animation: spin 1s linear infinite;
}}
@keyframes spin {{
  from {{ transform: rotate(0deg); }}
  to {{ transform: rotate(360deg); }}
}}

/* 편집 모드 */
.edit-toolbar {{
  display: flex;
  gap: 8px;
  margin-bottom: 12px;
}}
.edit-toolbar button {{
  padding: 8px 16px;
  border-radius: 8px;
  border: 1px solid var(--border);
  background: var(--card);
  color: var(--text);
  font-size: 13px;
  font-weight: 600;
  cursor: pointer;
}}
.edit-toolbar button.primary {{
  background: var(--accent);
  border-color: var(--accent);
  color: #fff;
}}
.edit-toolbar button.danger {{
  background: rgba(225,112,85,0.2);
  border-color: var(--red);
  color: var(--red);
}}
.holding-card .edit-actions {{
  display: none;
  gap: 6px;
  margin-top: 8px;
  padding-top: 8px;
  border-top: 1px solid var(--border);
}}
.edit-mode .holding-card .edit-actions {{ display: flex; }}
.edit-actions button {{
  padding: 6px 12px;
  border-radius: 6px;
  border: 1px solid var(--border);
  background: var(--card-hover);
  color: var(--text);
  font-size: 11px;
  cursor: pointer;
}}
.edit-actions button.del {{ color: var(--red); border-color: var(--red); }}

/* 모달 */
.modal-overlay {{
  display: none;
  position: fixed;
  top: 0; left: 0; right: 0; bottom: 0;
  background: rgba(0,0,0,0.7);
  z-index: 200;
  align-items: flex-end;
  justify-content: center;
}}
.modal-overlay.open {{ display: flex; }}
.modal {{
  background: var(--card);
  border-radius: 16px 16px 0 0;
  padding: 20px 16px 30px;
  width: 100%;
  max-width: 500px;
  max-height: 85vh;
  overflow-y: auto;
  animation: slideUp 0.3s ease;
}}
@keyframes slideUp {{
  from {{ transform: translateY(100%); }}
  to {{ transform: translateY(0); }}
}}
.modal h3 {{
  font-size: 16px;
  font-weight: 700;
  margin-bottom: 16px;
  display: flex;
  justify-content: space-between;
  align-items: center;
}}
.modal h3 button {{
  background: none;
  border: none;
  color: var(--text-sub);
  font-size: 20px;
  cursor: pointer;
}}
.form-group {{
  margin-bottom: 12px;
}}
.form-group label {{
  display: block;
  font-size: 11px;
  color: var(--text-sub);
  margin-bottom: 4px;
  font-weight: 600;
}}
.form-group input, .form-group select {{
  width: 100%;
  padding: 10px 12px;
  border-radius: 8px;
  border: 1px solid var(--border);
  background: var(--bg);
  color: var(--text);
  font-size: 14px;
  outline: none;
}}
.form-group input:focus, .form-group select:focus {{
  border-color: var(--accent);
}}
.form-row {{
  display: grid;
  grid-template-columns: 1fr 1fr;
  gap: 10px;
}}
.modal .btn-save {{
  width: 100%;
  padding: 12px;
  border-radius: 10px;
  border: none;
  background: var(--accent);
  color: #fff;
  font-size: 15px;
  font-weight: 600;
  cursor: pointer;
  margin-top: 8px;
}}
.ls-badge {{
  font-size: 10px;
  padding: 2px 6px;
  border-radius: 4px;
  background: rgba(253,203,110,0.2);
  color: var(--yellow);
  margin-left: 6px;
}}

/* 반응형 */
@media (min-width: 600px) {{
  .tab-content {{ padding: 20px; max-width: 600px; margin: 0 auto; }}
  .summary-grid {{ grid-template-columns: repeat(2, 1fr); }}
  .modal {{ border-radius: 16px; margin-bottom: 20px; }}
}}
</style>
</head>
<body>

<div class="header">
  <h1>
    My Portfolio
    <button id="refreshBtn" onclick="fetchPrices()" style="background:none;border:none;color:var(--accent-light);font-size:16px;cursor:pointer;padding:4px 8px;">&#8635;</button>
  </h1>
  <div class="date" id="updateDate"></div>
  <div id="priceStatus" style="font-size:11px;color:var(--yellow);margin-top:4px;display:none;"></div>
</div>

<nav class="tab-nav">
  <button class="tab-btn active" onclick="switchTab('dashboard')">대시보드</button>
  <button class="tab-btn" onclick="switchTab('holdings')">보유종목</button>
  <button class="tab-btn" onclick="switchTab('strategy')">투자전략</button>
  <button class="tab-btn" onclick="switchTab('watchlist')">워치리스트</button>
  <button class="tab-btn" onclick="switchTab('trades')">거래내역</button>
</nav>

<div id="tab-dashboard" class="tab-content active"></div>
<div id="tab-holdings" class="tab-content"></div>
<div id="tab-strategy" class="tab-content"></div>
<div id="tab-watchlist" class="tab-content"></div>
<div id="tab-trades" class="tab-content"></div>

<!-- 편집 모달 -->
<div class="modal-overlay" id="editModal">
  <div class="modal">
    <h3>
      <span id="modalTitle">종목 편집</span>
      <button onclick="closeModal()">&times;</button>
    </h3>
    <form id="editForm" onsubmit="saveHolding(event)">
      <input type="hidden" id="editIdx">
      <div class="form-group">
        <label>종목명</label>
        <input type="text" id="f_name" required>
      </div>
      <div class="form-row">
        <div class="form-group">
          <label>종목코드</label>
          <input type="text" id="f_ticker">
        </div>
        <div class="form-group">
          <label>시장</label>
          <select id="f_market">
            <option value="한국주식">한국주식</option>
            <option value="미국주식">미국주식</option>
            <option value="호주주식">호주주식</option>
          </select>
        </div>
      </div>
      <div class="form-row">
        <div class="form-group">
          <label>Tag</label>
          <input type="text" id="f_tag">
        </div>
        <div class="form-group">
          <label>업종</label>
          <input type="text" id="f_industry">
        </div>
      </div>
      <div class="form-row">
        <div class="form-group">
          <label>매입가</label>
          <input type="number" id="f_buyPrice" step="any">
        </div>
        <div class="form-group">
          <label>현재가</label>
          <input type="number" id="f_currentPrice" step="any">
        </div>
      </div>
      <div class="form-row">
        <div class="form-group">
          <label>주식수</label>
          <input type="number" id="f_shares" step="any">
        </div>
        <div class="form-group">
          <label>현재가치 (원)</label>
          <input type="number" id="f_currentValue" step="any">
        </div>
      </div>
      <button type="submit" class="btn-save">저장</button>
    </form>
  </div>
</div>

<script>
const ORIGINAL_DATA = {data_json};

// localStorage 기반 보유종목 관리
const LS_KEY = 'portfolio_holdings_edit';

function loadHoldings() {{
  const saved = localStorage.getItem(LS_KEY);
  if (saved) {{
    try {{ return JSON.parse(saved); }}
    catch(e) {{ return null; }}
  }}
  return null;
}}

function saveHoldingsToLS(holdings) {{
  localStorage.setItem(LS_KEY, JSON.stringify(holdings));
}}

function resetHoldings() {{
  if (confirm('편집한 내용을 모두 초기화하고 원본 데이터로 되돌릴까요?')) {{
    localStorage.removeItem(LS_KEY);
    DATA.holdings.holdings = JSON.parse(JSON.stringify(ORIGINAL_DATA.holdings.holdings));
    renderHoldings();
  }}
}}

// 데이터 초기화 (localStorage 우선)
const DATA = JSON.parse(JSON.stringify(ORIGINAL_DATA));
const savedHoldings = loadHoldings();
if (savedHoldings) {{
  DATA.holdings.holdings = savedHoldings;
}}

let editMode = false;

function toggleEditMode() {{
  editMode = !editMode;
  renderHoldings();
}}

function openAddModal() {{
  document.getElementById('modalTitle').textContent = '종목 추가';
  document.getElementById('editIdx').value = '-1';
  document.getElementById('f_name').value = '';
  document.getElementById('f_ticker').value = '';
  document.getElementById('f_market').value = '한국주식';
  document.getElementById('f_tag').value = '';
  document.getElementById('f_industry').value = '';
  document.getElementById('f_buyPrice').value = '';
  document.getElementById('f_currentPrice').value = '';
  document.getElementById('f_shares').value = '';
  document.getElementById('f_currentValue').value = '';
  document.getElementById('editModal').classList.add('open');
}}

function openEditModal(idx) {{
  const h = DATA.holdings.holdings[idx];
  document.getElementById('modalTitle').textContent = '종목 편집';
  document.getElementById('editIdx').value = idx;
  document.getElementById('f_name').value = h.name || '';
  document.getElementById('f_ticker').value = h.ticker || '';
  document.getElementById('f_market').value = h.market || '한국주식';
  document.getElementById('f_tag').value = h.tag || '';
  document.getElementById('f_industry').value = h.industry || '';
  document.getElementById('f_buyPrice').value = h.buyPrice || '';
  document.getElementById('f_currentPrice').value = h.currentPrice || '';
  document.getElementById('f_shares').value = h.shares || '';
  document.getElementById('f_currentValue').value = h.currentValue || '';
  document.getElementById('editModal').classList.add('open');
}}

function closeModal() {{
  document.getElementById('editModal').classList.remove('open');
}}

function saveHolding(e) {{
  e.preventDefault();
  const idx = parseInt(document.getElementById('editIdx').value);
  const formData = {{
    name: document.getElementById('f_name').value,
    ticker: document.getElementById('f_ticker').value,
    market: document.getElementById('f_market').value,
    tag: document.getElementById('f_tag').value,
    industry: document.getElementById('f_industry').value,
    buyPrice: parseFloat(document.getElementById('f_buyPrice').value) || null,
    currentPrice: parseFloat(document.getElementById('f_currentPrice').value) || null,
    shares: parseFloat(document.getElementById('f_shares').value) || null,
    currentValue: parseFloat(document.getElementById('f_currentValue').value) || null,
  }};

  // 수익률 자동 계산
  if (formData.buyPrice && formData.currentPrice) {{
    formData.returnPct = (formData.currentPrice - formData.buyPrice) / formData.buyPrice;
  }}
  // 현재가치 자동 계산 (없으면)
  if (!formData.currentValue && formData.currentPrice && formData.shares) {{
    formData.currentValue = formData.currentPrice * formData.shares;
  }}
  // 손익 계산
  if (formData.buyPrice && formData.currentPrice && formData.shares) {{
    formData.pnl = (formData.currentPrice - formData.buyPrice) * formData.shares;
  }}

  if (idx === -1) {{
    // 신규 추가
    formData.rank = null;
    formData.weight = null;
    formData.dailyChange = null;
    formData.dailyPnl = null;
    formData._edited = true;
    DATA.holdings.holdings.push(formData);
  }} else {{
    // 기존 수정 (기존 값 유지하면서 덮어쓰기)
    Object.assign(DATA.holdings.holdings[idx], formData);
    DATA.holdings.holdings[idx]._edited = true;
  }}

  saveHoldingsToLS(DATA.holdings.holdings);
  closeModal();
  renderHoldings();
}}

function deleteHolding(idx) {{
  const name = DATA.holdings.holdings[idx].name;
  if (confirm(name + ' 종목을 삭제할까요?')) {{
    DATA.holdings.holdings.splice(idx, 1);
    saveHoldingsToLS(DATA.holdings.holdings);
    renderHoldings();
  }}
}}

// Yahoo Finance 티커 변환
function toYahooTicker(ticker) {{
  if (!ticker) return null;
  if (ticker.startsWith('KOSDAQ:')) return ticker.replace('KOSDAQ:', '') + '.KQ';
  if (ticker.startsWith('KRX:')) return ticker.replace('KRX:', '') + '.KS';
  if (/^\\d{{6}}$/.test(ticker)) return ticker + '.KQ';
  if (ticker.startsWith('ASX:')) return ticker.replace('ASX:', '') + '.AX';
  return ticker;
}}

// 티커로 통화 결정 (시장 필드보다 티커 형식이 정확)
function getCurrencyFromTicker(ticker) {{
  if (!ticker) return 'KRW';
  if (ticker.endsWith('.KS') || ticker.endsWith('.KQ')) return 'KRW';
  if (ticker.endsWith('.AX')) return 'AUD';
  return 'USD';
}}

// 실시간 가격 가져오기
let fxRates = {{ usdkrw: 1484.57, audkrw: 1048.64 }};

async function fetchPrices() {{
  const btn = document.getElementById('refreshBtn');
  const status = document.getElementById('priceStatus');
  btn.classList.add('loading');
  status.style.display = 'block';
  status.textContent = '가격 업데이트 중...';
  status.style.color = 'var(--yellow)';

  try {{
    const holdings = DATA.holdings.holdings;
    const yahooTickers = holdings
      .map(h => toYahooTicker(h.ticker))
      .filter(Boolean);

    // 환율도 함께 요청
    const allSymbols = [...new Set([...yahooTickers, 'KRW=X', 'AUDKRW=X'])];
    const symbolStr = allSymbols.join(',');

    const apiUrl = 'https://query1.finance.yahoo.com/v7/finance/quote?symbols=' + symbolStr;
    // CORS 프록시 사용
    const proxyUrl = 'https://api.allorigins.win/raw?url=' + encodeURIComponent(apiUrl);

    const res = await fetch(proxyUrl, {{ signal: AbortSignal.timeout(15000) }});
    const json = await res.json();
    const quotes = json.quoteResponse?.result || [];

    if (quotes.length === 0) throw new Error('No data');

    // 가격 맵 생성
    const priceMap = {{}};
    quotes.forEach(q => {{
      if (q.symbol === 'KRW=X') {{ fxRates.usdkrw = q.regularMarketPrice; return; }}
      if (q.symbol === 'AUDKRW=X') {{ fxRates.audkrw = q.regularMarketPrice; return; }}
      priceMap[q.symbol] = q.regularMarketPrice;
    }});

    let updated = 0;
    holdings.forEach(h => {{
      const yt = toYahooTicker(h.ticker);
      if (!yt || priceMap[yt] == null) return;

      const newPrice = priceMap[yt];
      const currency = getCurrencyFromTicker(yt);
      let fxRate = 1;
      if (currency === 'USD') fxRate = fxRates.usdkrw;
      else if (currency === 'AUD') fxRate = fxRates.audkrw;

      h.currentPrice = newPrice;
      h.currentValue = newPrice * (h.shares || 0) * fxRate;
      if (h.buyPrice) {{
        h.returnPct = (newPrice - h.buyPrice) / h.buyPrice;
        h.pnl = (newPrice - h.buyPrice) * (h.shares || 0) * fxRate;
      }}
      updated++;
    }});

    // 시장별 요약 재계산
    recalcSummary();

    // localStorage에 저장
    saveHoldingsToLS(holdings);

    // 마지막 업데이트 시간 저장
    const now = new Date();
    localStorage.setItem('portfolio_price_updated', now.toISOString());

    // 렌더링
    renderHoldings();
    renderDashboard();

    status.textContent = updated + '종목 업데이트 완료 (' + now.toLocaleTimeString('ko-KR') + ')';
    status.style.color = 'var(--green)';
  }} catch(e) {{
    console.error('Price fetch error:', e);
    status.textContent = '가격 업데이트 실패: ' + e.message;
    status.style.color = 'var(--red)';
  }} finally {{
    btn.classList.remove('loading');
    setTimeout(() => {{ status.style.display = 'none'; }}, 5000);
  }}
}}

// 시장별 요약 재계산
function recalcSummary() {{
  const holdings = DATA.holdings.holdings;
  const marketTotals = {{}};

  holdings.forEach(h => {{
    const m = h.market;
    if (!m) return;
    if (!marketTotals[m]) marketTotals[m] = {{ value: 0, pnl: 0 }};
    marketTotals[m].value += h.currentValue || 0;
    marketTotals[m].pnl += h.pnl || 0;
  }});

  // 총 주식 가치
  let totalStockValue = 0;
  Object.values(marketTotals).forEach(t => totalStockValue += t.value);

  // 총 자산 = 주식 + 예수금
  const cash = DATA.holdings.summary.cash || 0;
  DATA.holdings.totalAssets = totalStockValue + cash;
  const total = DATA.holdings.totalAssets;

  // 시장별 요약 업데이트
  for (const [market, totals] of Object.entries(marketTotals)) {{
    if (DATA.holdings.summary[market]) {{
      const origCost = totals.value - totals.pnl;
      DATA.holdings.summary[market].value = totals.value;
      DATA.holdings.summary[market].weight = total > 0 ? totals.value / total : 0;
      DATA.holdings.summary[market].pnl = totals.pnl;
      DATA.holdings.summary[market].return = origCost > 0 ? totals.pnl / origCost : 0;
    }}
  }}
}}

// 유틸리티 함수
function formatKRW(val) {{
  if (val == null) return '-';
  if (val >= 100000000) return (val / 100000000).toFixed(1) + '억';
  if (val >= 10000) return (val / 10000).toFixed(0) + '만';
  return Math.round(val).toLocaleString();
}}

function formatPct(val) {{
  if (val == null) return '-';
  const pct = (val * 100).toFixed(1);
  return (val >= 0 ? '+' : '') + pct + '%';
}}

function pctClass(val) {{
  if (val == null) return '';
  return val >= 0 ? 'positive' : 'negative';
}}

function pctColor(val) {{
  if (val == null) return 'var(--text-sub)';
  return val >= 0 ? 'var(--green)' : 'var(--red)';
}}

// 탭 전환
function switchTab(tabId) {{
  document.querySelectorAll('.tab-content').forEach(t => t.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
  document.getElementById('tab-' + tabId).classList.add('active');
  event.target.classList.add('active');
}}

// 아코디언 토글
function toggleAccordion(el) {{
  el.closest('.accordion').classList.toggle('open');
}}

// 대시보드 렌더링
function renderDashboard() {{
  const h = DATA.holdings;
  const pv = DATA.portfolioValue;
  const totalAssets = h.totalAssets;
  const startVal = pv.startingValue;
  const totalReturn = startVal ? (totalAssets - startVal) / startVal : 0;
  const totalPnl = totalAssets - startVal;

  // 자산 배분 계산
  const cashWeight = h.summary.cash ? h.summary.cash / totalAssets : 0;
  const usWeight = h.summary['미국주식'] ? h.summary['미국주식'].weight : 0;
  const krWeight = h.summary['한국주식'] ? h.summary['한국주식'].weight : 0;
  const auWeight = h.summary['호주주식'] ? h.summary['호주주식'].weight : 0;

  document.getElementById('updateDate').textContent = '최종 업데이트: ' + (pv.months.length > 0 ? pv.months[pv.months.length - 1].date : '-');

  let html = `
    <div class="summary-grid">
      <div class="summary-card full">
        <div class="label">총 자산</div>
        <div class="value" style="font-size: 28px;">₩${{formatKRW(totalAssets)}}</div>
        <div class="sub">
          총 수익: <span style="color:${{pctColor(totalReturn)}}">₩${{formatKRW(totalPnl)}} (${{formatPct(totalReturn)}})</span>
        </div>
      </div>
      <div class="summary-card">
        <div class="label">한국주식</div>
        <div class="value ${{pctClass(h.summary['한국주식']?.return)}}">
          ${{formatPct(h.summary['한국주식']?.return)}}
        </div>
        <div class="sub">₩${{formatKRW(h.summary['한국주식']?.value)}}</div>
      </div>
      <div class="summary-card">
        <div class="label">미국주식</div>
        <div class="value ${{pctClass(h.summary['미국주식']?.return)}}">
          ${{formatPct(h.summary['미국주식']?.return)}}
        </div>
        <div class="sub">₩${{formatKRW(h.summary['미국주식']?.value)}}</div>
      </div>
      <div class="summary-card">
        <div class="label">호주주식</div>
        <div class="value ${{pctClass(h.summary['호주주식']?.return)}}">
          ${{formatPct(h.summary['호주주식']?.return)}}
        </div>
        <div class="sub">₩${{formatKRW(h.summary['호주주식']?.value)}}</div>
      </div>
      <div class="summary-card">
        <div class="label">예수금</div>
        <div class="value" style="color: var(--text);">₩${{formatKRW(h.summary.cash)}}</div>
        <div class="sub">${{(cashWeight * 100).toFixed(1)}}%</div>
      </div>
    </div>

    <div class="chart-container">
      <h3>자산 배분</h3>
      <canvas id="allocationChart"></canvas>
      <div class="alloc-grid">
        <div class="alloc-item">
          <div class="alloc-dot" style="background: #6c5ce7;"></div>
          <div class="alloc-info">
            <div class="alloc-label">미국주식</div>
            <div class="alloc-value">${{(usWeight * 100).toFixed(1)}}%</div>
          </div>
        </div>
        <div class="alloc-item">
          <div class="alloc-dot" style="background: #00b894;"></div>
          <div class="alloc-info">
            <div class="alloc-label">한국주식</div>
            <div class="alloc-value">${{(krWeight * 100).toFixed(1)}}%</div>
          </div>
        </div>
        <div class="alloc-item">
          <div class="alloc-dot" style="background: #fdcb6e;"></div>
          <div class="alloc-info">
            <div class="alloc-label">호주주식</div>
            <div class="alloc-value">${{(auWeight * 100).toFixed(1)}}%</div>
          </div>
        </div>
        <div class="alloc-item">
          <div class="alloc-dot" style="background: #636e72;"></div>
          <div class="alloc-info">
            <div class="alloc-label">예수금</div>
            <div class="alloc-value">${{(cashWeight * 100).toFixed(1)}}%</div>
          </div>
        </div>
      </div>
    </div>

    <div class="chart-container">
      <h3>월별 누적 수익률 (Indexed)</h3>
      <canvas id="returnChart"></canvas>
    </div>
  `;

  document.getElementById('tab-dashboard').innerHTML = html;

  // 자산배분 도넛 차트
  new Chart(document.getElementById('allocationChart'), {{
    type: 'doughnut',
    data: {{
      labels: ['미국주식', '한국주식', '호주주식', '예수금'],
      datasets: [{{
        data: [usWeight, krWeight, auWeight, cashWeight],
        backgroundColor: ['#6c5ce7', '#00b894', '#fdcb6e', '#636e72'],
        borderWidth: 0,
        borderRadius: 4,
      }}]
    }},
    options: {{
      responsive: true,
      cutout: '65%',
      plugins: {{
        legend: {{ display: false }},
        tooltip: {{
          callbacks: {{
            label: ctx => ctx.label + ': ' + (ctx.parsed * 100).toFixed(1) + '%'
          }}
        }}
      }}
    }}
  }});

  // 월별 수익률 차트
  const months = pv.months.filter(m => m.portfolioIndexed != null);
  const labels = months.map(m => {{
    const d = m.date;
    if (!d) return '';
    const parts = d.split('-');
    return parts[0].slice(2) + '/' + parts[1];
  }});

  new Chart(document.getElementById('returnChart'), {{
    type: 'line',
    data: {{
      labels: labels,
      datasets: [
        {{
          label: 'My Portfolio',
          data: months.map(m => m.portfolioIndexed ? ((m.portfolioIndexed - 1) * 100).toFixed(1) : null),
          borderColor: '#6c5ce7',
          backgroundColor: 'rgba(108,92,231,0.1)',
          borderWidth: 2,
          pointRadius: 3,
          fill: true,
          tension: 0.3,
        }},
        {{
          label: 'S&P 500',
          data: months.map(m => m.sp500Indexed ? ((m.sp500Indexed - 1) * 100).toFixed(1) : null),
          borderColor: '#74b9ff',
          borderWidth: 1.5,
          pointRadius: 2,
          borderDash: [4, 4],
          tension: 0.3,
        }},
        {{
          label: 'KOSPI',
          data: months.map(m => m.kospiIndexed ? ((m.kospiIndexed - 1) * 100).toFixed(1) : null),
          borderColor: '#00b894',
          borderWidth: 1.5,
          pointRadius: 2,
          borderDash: [4, 4],
          tension: 0.3,
        }}
      ]
    }},
    options: {{
      responsive: true,
      interaction: {{ mode: 'index', intersect: false }},
      scales: {{
        y: {{
          ticks: {{
            callback: v => v + '%',
            color: '#8b8fa3',
            font: {{ size: 10 }}
          }},
          grid: {{ color: 'rgba(42,45,58,0.5)' }},
        }},
        x: {{
          ticks: {{ color: '#8b8fa3', font: {{ size: 10 }} }},
          grid: {{ display: false }}
        }}
      }},
      plugins: {{
        legend: {{
          labels: {{ color: '#e4e6eb', font: {{ size: 11 }}, boxWidth: 12 }}
        }}
      }}
    }}
  }});
}}

// 보유종목 렌더링
function renderHoldings() {{
  const holdings = DATA.holdings.holdings;
  const hasEdits = localStorage.getItem(LS_KEY) != null;

  // 시장별 그룹핑 (원본 인덱스 보존)
  const groups = {{}};
  holdings.forEach((h, i) => {{
    const m = h.market || '기타';
    if (!groups[m]) groups[m] = [];
    groups[m].push({{ ...h, _idx: i }});
  }});

  const marketOrder = ['한국주식', '미국주식', '호주주식'];

  let html = `
    <div class="edit-toolbar">
      <button class="${{editMode ? 'danger' : 'primary'}}" onclick="toggleEditMode()">
        ${{editMode ? '편집 완료' : '편집'}}
      </button>
      ${{editMode ? '<button class="primary" onclick="openAddModal()">+ 종목 추가</button>' : ''}}
      ${{hasEdits ? '<button class="danger" onclick="resetHoldings()">초기화</button>' : ''}}
      ${{hasEdits ? '<span class="ls-badge" style="align-self:center;">편집됨</span>' : ''}}
    </div>`;

  const container = editMode ? ' edit-mode' : '';

  html += `<div class="${{container}}">`;

  marketOrder.forEach(market => {{
    const items = groups[market];
    if (!items) return;

    const summary = DATA.holdings.summary[market];
    html += `
      <div class="market-group-header">
        <span>${{market}} (${{items.length}}종목)</span>
        <span class="market-return" style="color:${{pctColor(summary?.return)}}">
          ${{formatPct(summary?.return)}}
        </span>
      </div>`;

    items.sort((a, b) => (b.currentValue || 0) - (a.currentValue || 0));
    items.forEach(h => {{
      const editBadge = h._edited ? '<span class="ls-badge">수정</span>' : '';
      html += `
        <div class="holding-card">
          <div class="holding-header">
            <div class="holding-name">${{h.name}}${{editBadge}}</div>
            <div class="holding-return" style="color:${{pctColor(h.returnPct)}}">
              ${{formatPct(h.returnPct)}}
            </div>
          </div>
          <div class="holding-meta">
            <span class="tag market">${{h.ticker}}</span>
            <span class="tag">${{h.tag || h.industry || ''}}</span>
          </div>
          <div class="holding-details">
            <span>
              <span class="detail-label">현재가치</span>
              <span class="detail-value">₩${{formatKRW(h.currentValue)}}</span>
            </span>
            <span>
              <span class="detail-label">비중</span>
              <span class="detail-value">${{h.weight ? (h.weight * 100).toFixed(1) + '%' : '-'}}</span>
            </span>
            <span>
              <span class="detail-label">손익</span>
              <span class="detail-value" style="color:${{pctColor(h.pnl)}}">₩${{formatKRW(h.pnl)}}</span>
            </span>
            <span>
              <span class="detail-label">일일등락</span>
              <span class="detail-value" style="color:${{pctColor(h.dailyChange)}}">
                ${{h.dailyChange != null ? formatPct(h.dailyChange) : '-'}}
              </span>
            </span>
          </div>
          <div class="edit-actions">
            <button onclick="openEditModal(${{h._idx}})">수정</button>
            <button class="del" onclick="deleteHolding(${{h._idx}})">삭제</button>
          </div>
        </div>`;
    }});
  }});

  html += '</div>';
  document.getElementById('tab-holdings').innerHTML = html;
}}

// 투자전략 렌더링
function renderStrategy() {{
  const s = DATA.strategy;
  const sectorColors = ['#6c5ce7', '#00b894', '#fdcb6e', '#74b9ff'];

  let html = `
    <div class="section-title">자산배분 전략</div>
    <div class="strategy-card">
      <h4>전략 시점: ${{s.strategyDate || '-'}}</h4>
      <p>${{s.assetAllocation.replace(/\\n/g, '<br>')}}</p>
    </div>

    <div class="section-title">섹터별 투자전략</div>
    <div class="sector-bar">
      ${{s.sectors.map((sec, i) => `
        <div style="width:${{(sec.weight || 0) * 100}}%; background:${{sectorColors[i]}}">
          ${{sec.name}}<br>${{((sec.weight || 0) * 100)}}%
        </div>
      `).join('')}}
    </div>
    <div class="strategy-card">
      ${{s.sectors.map((sec, i) => `
        <div style="display:flex; justify-content:space-between; padding:6px 0; border-bottom:1px solid var(--border);">
          <span style="display:flex; align-items:center; gap:6px;">
            <span style="width:10px;height:10px;border-radius:50%;background:${{sectorColors[i]}};display:inline-block;"></span>
            ${{sec.name}}
          </span>
          <span style="color:var(--text-sub);">비중 ${{((sec.weight || 0) * 100)}}% / 목표수익률 ${{((sec.targetReturn || 0) * 100)}}%</span>
        </div>
      `).join('')}}
    </div>

    <div class="strategy-card">
      <h4>투자 전략 상세</h4>
      <p>${{s.investmentStrategy.replace(/\\n/g, '<br>')}}</p>
    </div>
  `;

  if (s.strategyMemo) {{
    html += `
      <div class="strategy-card">
        <h4>전략 메모</h4>
        <p>${{String(s.strategyMemo).replace(/\\n/g, '<br>')}}</p>
      </div>`;
  }}

  html += '<div class="section-title">종목별 투자판단</div>';

  s.judgements.forEach(j => {{
    let decisionClass = 'hold';
    const dec = (j.decision || '').toLowerCase();
    if (dec.includes('매수') || dec === 'buy') decisionClass = 'buy';
    else if (dec.includes('매도') || dec === 'sale' || dec === 'sell') decisionClass = 'sell';
    else if (dec.includes('not')) decisionClass = 'notinvest';

    const resultHtml = j.result ? `
      <div class="field">
        <div class="field-label">사후 결과</div>
        <div class="field-value">${{j.result.replace(/\\n/g, '<br>')}}</div>
      </div>` : '';
    const returnHtml = j.resultReturn != null ? `
      <div class="field">
        <div class="field-label">결과 수익률</div>
        <div class="field-value" style="color:${{pctColor(j.resultReturn)}}; font-size:14px; font-weight:600;">
          ${{formatPct(j.resultReturn)}}
        </div>
      </div>` : '';

    html += `
      <div class="accordion">
        <div class="accordion-header" onclick="toggleAccordion(this)">
          <span class="acc-date">${{j.date || '-'}}</span>
          <span class="acc-title">${{j.name}}</span>
          <span class="acc-decision ${{decisionClass}}">${{j.decision || '-'}}</span>
          <span class="arrow">▼</span>
        </div>
        <div class="accordion-body">
          <div class="field">
            <div class="field-label">카테고리</div>
            <div class="field-value">${{j.category || '-'}} / ${{j.industry || '-'}}</div>
          </div>
          ${{j.thesis ? `
          <div class="field">
            <div class="field-label">Investment Thesis</div>
            <div class="field-value">${{j.thesis.replace(/\\n/g, '<br>')}}</div>
          </div>` : ''}}
          ${{j.logic ? `
          <div class="field">
            <div class="field-label">판단 논리</div>
            <div class="field-value">${{j.logic.replace(/\\n/g, '<br>')}}</div>
          </div>` : ''}}
          ${{resultHtml}}
          ${{returnHtml}}
        </div>
      </div>`;
  }});

  document.getElementById('tab-strategy').innerHTML = html;
}}

// 워치리스트 렌더링
function renderWatchlist() {{
  const groups = DATA.watchlist;
  const groupOrder = ['제로투원', 'SJL 파트너스', '지인 픽(주경 / 우영 / 규현)', '올바른 / 바바리안', '발전기 섹터', '데이터센터 / 통신', '기타'];
  let html = '';

  // 정렬: 정의된 순서 먼저, 나머지는 이후
  const allGroups = [...new Set([...groupOrder, ...Object.keys(groups)])];

  allGroups.forEach(groupName => {{
    const items = groups[groupName];
    if (!items || items.length === 0) return;

    // 그룹명 단순화
    let displayName = groupName;
    if (groupName.includes('에스티어') || groupName.includes('수영님') || groupName.includes('재욱님') || groupName.includes('동하님')) {{
      displayName = '지인 픽';
    }}

    html += `<div class="wl-group">
      <div class="wl-group-title">${{displayName}} (${{items.length}})</div>`;

    items.forEach(w => {{
      const changeColor = pctColor(w.dailyChange);
      html += `
        <div class="wl-card">
          <div class="wl-card-header">
            <div class="wl-card-name">${{w.name}}</div>
            <div class="wl-card-change" style="color:${{changeColor}}">
              ${{w.dailyChange != null ? formatPct(w.dailyChange) : '-'}}
            </div>
          </div>
          <div class="wl-card-meta">
            <span class="tag market">${{w.ticker || '-'}}</span>
            <span class="tag">${{w.tag || w.industry || '-'}}</span>
            ${{w.market ? '<span>' + w.market + '</span>' : ''}}
            ${{w.marketCap ? '<span>시총 ' + (typeof w.marketCap === 'number' ? w.marketCap.toFixed(0) + 'B' : w.marketCap) + '</span>' : ''}}
          </div>
          ${{w.idea ? '<div style="margin-top:8px;font-size:11px;color:var(--text-sub);line-height:1.5;">' + w.idea.replace(/\\n/g, '<br>') + '</div>' : ''}}
        </div>`;
    }});

    html += '</div>';
  }});

  document.getElementById('tab-watchlist').innerHTML = html;
}}

// 거래내역 렌더링
function renderTrades() {{
  const trades = DATA.trades;
  let html = '<div class="section-title">최근 거래내역</div>';

  trades.forEach(t => {{
    const isSell = t.action.toLowerCase().includes('sale') || t.action.toLowerCase().includes('매도');
    const actionClass = isSell ? 'sell' : 'buy';
    const actionLabel = isSell ? '매도' : t.action;

    html += `
      <div class="trade-card">
        <div class="trade-header">
          <div class="trade-name">${{t.name}}</div>
          <span class="trade-action ${{actionClass}}">${{actionLabel}}</span>
        </div>
        <div style="font-size:11px; color:var(--text-sub); margin-bottom:8px;">
          ${{t.date || '-'}} · ${{t.ticker}} · ${{t.market}}
        </div>
        <div class="trade-details">
          <div>
            <div class="td-label">매입가</div>
            <div class="td-value">${{t.buyPrice != null ? t.buyPrice.toLocaleString() : '-'}}</div>
          </div>
          <div>
            <div class="td-label">${{isSell ? '매도가' : '현재가치'}}</div>
            <div class="td-value">${{isSell ? (t.sellPrice != null ? t.sellPrice.toLocaleString() : '-') : '₩' + formatKRW(t.currentValue)}}</div>
          </div>
          <div>
            <div class="td-label">수익률</div>
            <div class="td-value" style="color:${{pctColor(t.returnPct)}}">${{formatPct(t.returnPct)}}</div>
          </div>
        </div>
        ${{t.rationale ? '<div class="trade-rationale">' + t.rationale + '</div>' : ''}}
      </div>`;
  }});

  document.getElementById('tab-trades').innerHTML = html;
}}

// 초기 렌더링
renderDashboard();
renderHoldings();
renderStrategy();
renderWatchlist();
renderTrades();

// 페이지 로드 시 자동 가격 업데이트
const lastUpdate = localStorage.getItem('portfolio_price_updated');
if (lastUpdate) {{
  const elapsed = Date.now() - new Date(lastUpdate).getTime();
  // 마지막 업데이트가 10분 이상 지났으면 자동 갱신
  if (elapsed > 10 * 60 * 1000) {{
    fetchPrices();
  }} else {{
    const t = new Date(lastUpdate);
    const el = document.getElementById('priceStatus');
    el.style.display = 'block';
    el.textContent = '마지막 가격 업데이트: ' + t.toLocaleTimeString('ko-KR');
    el.style.color = 'var(--text-sub)';
    setTimeout(() => {{ el.style.display = 'none'; }}, 3000);
  }}
}} else {{
  // 첫 방문이면 바로 가격 가져오기
  fetchPrices();
}}
</script>
</body>
</html>'''

    return html


def main():
    print('엑셀 파일 읽는 중...')
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    print('Holdings 데이터 추출...')
    holdings_data = read_holdings(wb)

    print('Portfolio Value 데이터 추출...')
    pv_data = read_portfolio_value(wb)

    print('Investment Judgement 데이터 추출...')
    strategy_data = read_investment_judgement(wb)

    print('Watchlist 데이터 추출...')
    watchlist_data = read_watchlist(wb)

    print('Trades 데이터 추출...')
    trades_data = read_trades(wb)

    data = {
        'holdings': holdings_data,
        'portfolioValue': pv_data,
        'strategy': strategy_data,
        'watchlist': watchlist_data,
        'trades': trades_data,
    }

    print('HTML 생성 중...')
    html = generate_html(data)

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        f.write(html)

    print(f'완료! {OUTPUT_PATH}')
    print(f'파일 크기: {os.path.getsize(OUTPUT_PATH) / 1024:.1f} KB')


if __name__ == '__main__':
    main()
